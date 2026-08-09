import os
import glob
import re
import json
from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import gspread
import pandas as pd
import gdown
from openpyxl import load_workbook

app = Flask(__name__)

# ดึงค่า LINE Token และ Secret จาก Environment Variables บน Render
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get('LINE_CHANNEL_ACCESS_TOKEN')
LINE_CHANNEL_SECRET = os.environ.get('LINE_CHANNEL_SECRET')

line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ใช้ ID โฟลเดอร์ My Stock เพื่อดึงไฟล์ล่าสุดอัตโนมัติ
DRIVE_FOLDER_ID = "19DLipG-4_C0qWTOsFGXyJWfhLsNvR4V8"
GSHEET_URL = "https://docs.google.com/spreadsheets/d/1ngb3u6xzE6m0QSre1gTwFxm0_hElAavyPGKkOHY98Vc/edit?gid=0#gid=0"

HEADER_ROW = 4
CODE_B_INDEX = 1
CODE_C_INDEX = 2
DESC_COL_INDEX = 3     
NEW_COL_INDEX = 12
OLD_COL_INDEX = 13
BALANCE_COL_INDEX = 63

def clean_num(val):
    if pd.isna(val) or val is None: return 0.0
    s = str(val).replace(',', '').strip()
    if s in ["-", "_", "", "nan", "None"]: return 0.0
    try: return float(s)
    except: return 0.0

def normalize_code(code_str):
    if not code_str: return ""
    return re.sub(r'[^A-Z0-9]', '', str(code_str).strip().upper())

def get_google_credentials():
    google_creds_json = os.environ.get('GOOGLE_CREDENTIALS_JSON')
    if google_creds_json:
        creds_dict = json.loads(google_creds_json)
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n")
        return Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    else:
        return Credentials.from_service_account_file("credentials.json", scopes=SCOPES)

def get_latest_file_ids_from_folder(creds, folder_id, count=2):
    """ ดึงเฉพาะ File ID ของไฟล์ Excel 2 ไฟล์ล่าสุดจากโฟลเดอร์ My Stock อัตโนมัติ """
    service = build('drive', 'v3', credentials=creds)
    query = f"'{folder_id}' in parents and (mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or mimeType='application/vnd.ms-excel') and trashed=false"
    results = service.files().list(q=query, orderBy="modifiedTime desc", pageSize=count, fields="files(id, name)").execute()
    items = results.get('files', [])
    return [item['id'] for item in items]

def process_order_and_get_summary(user_msg):
    creds = get_google_credentials()
    client = gspread.authorize(creds)
    spreadsheet = client.open_by_url(GSHEET_URL)

    # 1. บันทึกข้อความที่ส่งมาลงใน Input_Order
    ws_input = spreadsheet.worksheet("Input_Order")
    ws_input.clear()
    
    lines = user_msg.strip().split('\n')
    input_data = [["Code", "Qty"]]
    for line in lines:
        parts = line.strip().split()
        if len(parts) >= 2:
            input_data.append([parts[0], parts[1]])
        elif len(parts) == 1:
            input_data.append([parts[0], "1"])

    ws_input.update(input_data, 'A1')

    # 2. ลบไฟล์เก่าทิ้งและดาวน์โหลดไฟล์ใหม่จาก Drive
    for f in glob.glob("*.xlsx"):
        try: os.remove(f)
        except: pass

    file_ids = get_latest_file_ids_from_folder(creds, DRIVE_FOLDER_ID, count=2)
    if not file_ids:
        return "❌ ไม่พบไฟล์ Excel ในโฟลเดอร์ My Stock"

    for file_id in file_ids:
        gdown.download(id=file_id, quiet=False)

    all_xlsx = [f for f in glob.glob("*.xlsx") if not os.path.basename(f).startswith("~$")]
    global_code_map = {}
    combined_headers = []

    # ใช้ openpyxl โหลดแบบไม่กินแรม แล้วแปลงเป็น df_raw เพื่อคงโครงสร้างเดิมของคุณไว้ทั้งหมด
    for file_path in all_xlsx:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()
        
        df_raw = pd.DataFrame(rows)
        num_cols = df_raw.shape[1]
        
        for r in range(HEADER_ROW + 1, len(df_raw)):
            for col_idx in [CODE_B_INDEX, CODE_C_INDEX]:
                c_val = df_raw.iloc[r, col_idx] if col_idx < num_cols else None
                if pd.notna(c_val):
                    norm_c = normalize_code(c_val)
                    if norm_c and norm_c not in ["NAN", "NONE", "0", "CODE", "รหัสสินค้า", "รหัส", "NO"]:
                        if norm_c not in global_code_map:
                            global_code_map[norm_c] = (df_raw, r)
                            
        for c in range(num_cols):
            txts = [str(df_raw.iloc[r, c]).strip() for r in range(HEADER_ROW, min(HEADER_ROW + 3, len(df_raw))) if r < len(df_raw) and pd.notna(df_raw.iloc[r, c])]
            combined_headers.append(" ".join(txts))

    df_input = pd.DataFrame(ws_input.get_all_records())
    excluded_kw = ["no", "code", "รายการ", "order", "category", "weight", "quantity", "qty", "on hand", "balance", "ขาด", "old", "new", "broken", "po", "repair", "total", "dift"]

    report_items = []
    input_code_col = df_input.columns[0]
    input_qty_col = df_input.columns[1]

    for _, row in df_input.iterrows():
        raw_code = str(row[input_code_col]).strip()
        norm_input = normalize_code(raw_code)
        qty_needed = clean_num(row[input_qty_col])
        
        match_data = global_code_map.get(norm_input)
        if match_data is not None:
            df_target, target_row = match_data
            balance = clean_num(df_target.iloc[target_row, BALANCE_COL_INDEX] if BALANCE_COL_INDEX < df_target.shape[1] else 0)
            shortage = qty_needed if balance < 0 else max(0, qty_needed - balance)
            
            b_val = df_target.iloc[target_row, CODE_B_INDEX] if CODE_B_INDEX < df_target.shape[1] else raw_code
            desc_val = df_target.iloc[target_row, DESC_COL_INDEX] if DESC_COL_INDEX < df_target.shape[1] else ""

            report_items.append({
                'code': str(b_val), 
                'desc': str(desc_val),
                'shortage': "มีของ" if shortage == 0 else int(shortage),
                'balance': int(balance)
            })
        else:
            report_items.append({
                'code': raw_code, 
                'desc': "ไม่พบรหัส",
                'shortage': int(qty_needed),
                'balance': "-"
            })

    # 3. สร้างข้อความสรุปผลส่งกลับไปที่หน้าจอ LINE ทันที
    summary_text = "📊 รายงานสรุปสต็อก:\n"
    for item in report_items:
        summary_text += f"- {item['code']} ({item['desc']}): คงเหลือ {item['balance']} | ขาด {item['shortage']}\n"

    # อัปเดตลงชีต Summary ด้วย
    header = ["รหัสสินค้า", "รายการ", "สต็อก Balance", "ขาด"]
    table_data = [header] + [[i['code'], i['desc'], i['balance'], i['shortage']] for i in report_items]
    ws_summary = spreadsheet.worksheet("Summary") if "Summary" in [w.title for w in spreadsheet.worksheets()] else spreadsheet.add_worksheet(title="Summary", rows="100", cols="30")
    ws_summary.clear()
    ws_summary.update(table_data, 'A1')

    return summary_text

@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers['X-Line-Signature']
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    return 'OK'

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    user_msg = event.message.text.strip()
    try:
        reply_text = process_order_and_get_summary(user_msg)
    except Exception as e:
        reply_text = f"❌ เกิดข้อผิดพลาด: {str(e)}"

    line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply_text))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
